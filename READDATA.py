import openpyxl

path= "D:\job prep\sqa\DATADRIVEN.xlsx"
workbook= openpyxl.load_workbook(path) #load the workbook from the path that is defined above
sheet= workbook.active  # or workbook.get_sheet_by_name("sheet1")

rows= sheet.max_row
col= sheet.max_column

print(rows)
print(col)

for r in range(1,rows+1):
    for c in range(1, col+1):
        print(sheet.cell(row=r, column=c).value, end= "  ")
    print()    